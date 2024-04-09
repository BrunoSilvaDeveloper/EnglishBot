import telebot
from telebot import types
import random
from openpyxl import load_workbook


CHAVE_API = "Sua chave API"

bot = telebot.TeleBot(CHAVE_API)

def responder(id, resposta, buttons, qtd):
    btn = []
    markup = types.InlineKeyboardMarkup(row_width=qtd)
    for bt in buttons:
        button = types.InlineKeyboardButton(bt[0], callback_data=bt[1])
        btn.append(button)

    markup.add(*btn)

    bot.send_message(id, resposta, reply_markup=markup)   

def responder_sem_button(id, resposta):
    bot.send_message(id, resposta)   

def gerar_numero():
    probabilidade = random.random() 
    
    if probabilidade < 0.4:
        return random.randint(0, 100) 
    elif probabilidade < 0.7:
        return random.randint(101, 500) 
    elif probabilidade < 0.9:
        return random.randint(501, 3000) 
    else:
        return random.randint(3001, 99999)
    
def verificar_dezena(numero):
    planilha = load_workbook('..\DataBase\Aprendendo.xlsx')
    aba_ativa = planilha['Numeros']

    if numero >= 0 and numero < 21:
        cel = f'A{numero+1}'
        return aba_ativa[cel].value

    elif numero == 30:
        cel = aba_ativa['A22'].value
        return cel

    elif numero == 40:
        cel = aba_ativa['A23'].value
        return cel

    elif numero == 50:
        cel = aba_ativa['A24'].value
        return cel

    elif numero == 60:
        cel = aba_ativa['A25'].value
        return cel

    elif numero == 70:
        cel = aba_ativa['A26'].value
        return cel

    elif numero == 80:
        cel = aba_ativa['A27'].value
        return cel

    elif numero == 90:
        cel = aba_ativa['A28'].value
        return cel
        
    elif numero > 20 and numero < 30:
        ab1 = aba_ativa['A21'].value
        ab2 = f'A{numero-19}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 30 and numero < 40:
        ab1 = aba_ativa['A22'].value
        ab2 = f'A{numero-29}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 40 and numero < 50:
        ab1 = aba_ativa['A23'].value
        ab2 = f'A{numero-39}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 50 and numero < 60:
        ab1 = aba_ativa['A24'].value
        ab2 = f'A{numero-49}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 60 and numero < 70:
        ab1 = aba_ativa['A25'].value
        ab2 = f'A{numero-59}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 70 and numero < 80:
        ab1 = aba_ativa['A26'].value
        ab2 = f'A{numero-69}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 80 and numero < 90:
        ab1 = aba_ativa['A27'].value
        ab2 = f'A{numero-79}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel

    elif numero > 90 and numero < 100:
        ab1 = aba_ativa['A28'].value
        ab2 = f'A{numero-89}'
        ab21 = aba_ativa[ab2].value
        cel = f'{ab1} {ab21}'
        return cel
    
def verificar_centena(numero):

    planilha = load_workbook('..\DataBase\Aprendendo.xlsx')
    aba_ativa = planilha['Numeros']

    if numero == 100:
        ab1 = aba_ativa['A2'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 200:
        ab1 = aba_ativa['A3'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 300:
        ab1 = aba_ativa['A4'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 400:
        ab1 = aba_ativa['A5'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 500:
        ab1 = aba_ativa['A6'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 600:
        ab1 = aba_ativa['A7'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel 
    
    if numero == 700:
        ab1 = aba_ativa['A8'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 800:
        ab1 = aba_ativa['A9'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
    
    if numero == 900:
        ab1 = aba_ativa['A10'].value
        ab2 = aba_ativa['A29'].value
        cel = f'{ab1} {ab2}'
        return cel
        
    elif numero > 100 and numero < 200:
        ab1 = aba_ativa['A2'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-100)
        cel = f'{ab1} {ab2} {ab3}'
        return cel

    elif numero > 200 and numero < 300:
        ab1 = aba_ativa['A3'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-200)
        cel = f'{ab1} {ab2} {ab3}'
        return cel
    
    elif numero > 300 and numero < 400:
        ab1 = aba_ativa['A4'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-300)
        cel = f'{ab1} {ab2} {ab3}'
        return cel

    elif numero > 400 and numero < 500:
        ab1 = aba_ativa['A5'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-400)
        cel = f'{ab1} {ab2} {ab3}'
        return cel
    
    elif numero > 500 and numero < 600:
        ab1 = aba_ativa['A6'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-500)
        cel = f'{ab1} {ab2} {ab3}'
        return cel
    
    elif numero > 600 and numero < 700:
        ab1 = aba_ativa['A7'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-600)
        cel = f'{ab1} {ab2} {ab3}'
        return cel
    
    elif numero > 700 and numero < 800:
        ab1 = aba_ativa['A8'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-700)
        cel = f'{ab1} {ab2} {ab3}'
        return cel

    elif numero > 800 and numero < 900:
        ab1 = aba_ativa['A9'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-800)
        cel = f'{ab1} {ab2} {ab3}'
        return cel

    elif numero > 900 and numero < 1000:
        ab1 = aba_ativa['A10'].value
        ab2 = aba_ativa['A29'].value
        ab3 = verificar_dezena(numero-900)
        cel = f'{ab1} {ab2} {ab3}'
        return cel

def verificar_milhar(numero):
    n = str(numero)
    if len(n) == 4:
        milhar = int(n[0])
        centena = int(n[-3])*100
        dezena = int(n[-2]+n[-1])
    elif len(n) == 5:
        milhar = int(n[0]+n[1])
        centena = int(n[-3])*100
        dezena = int(n[-2]+n[-1])

    if numero > 999 and numero < 100000:
        if centena == 0:
            ab1 = verificar_dezena(milhar)
            ab3 = verificar_dezena(dezena)
            cel = f'{ab1} thousand {ab3}'
            return cel
        else:
            ab1 = verificar_dezena(milhar)
            ab2 = verificar_centena(centena)
            ab3 = verificar_dezena(dezena)
            cel = f'{ab1} thousand {ab2} {ab3}'
            return cel
        
def decidir_verificacao_number(numero):
    if numero < 100:
        cel = verificar_dezena(numero).capitalize()
        return cel

    elif numero >=100 and numero < 1000:
        cel = verificar_centena(numero).capitalize()
        return cel

    elif numero >=1000:
        cel = verificar_milhar(numero).capitalize()
        return cel

def Numbers(mensagem, user):
    id = user.get_id()
    number = user.get_number()

    if mensagem == '/Numbers':
        resposta = f'Você quer aprender sobre números?! 👋 \n\nSelecione o que deseja!'
        responder(id, resposta, [['Conteúdo sobre Numbers', '/ConteudoNumbers'], ['Exibir Number', '/ExibirNumber']], 1)
        

    elif mensagem == '/ConteudoNumbers':
        planilha = load_workbook('..\DataBase\Aprendendo.xlsx')
        aba_ativa = planilha['Numeros']
        resposta = aba_ativa['D1'].value
        responder(id, resposta, [['Exibir Number', '/ExibirNumber'], ['Continuar','/OK']], 1)

    elif mensagem == '/ExibirNumber':
        numero = gerar_numero()
        user.set_number(numero)
        resposta = f'O número escolhido foi {numero} '
        responder(id, resposta, [['Escrever por extenso', '/ExtensoNumbers'], ['Exibir Número Extenso', '/ExibirNumberExtenso'], ['Exibir Number', '/ExibirNumber'], ['Continuar','/OK']], 1)
    
    elif mensagem == '/ExtensoNumbers':
        user.set_ultimoComando('/ExtensoNumbers')
        resposta = f'Digite o número informado por extenso: '
        responder_sem_button(id, resposta)

    elif mensagem == '/ExibirNumberExtenso':
        extenso = decidir_verificacao_number(int(number))
        resposta = f'\nO número {number} por extenso é: \n\n{extenso}'
        responder(id, resposta, [['Exibir Number', '/ExibirNumber'], ['Continuar','/OK']], 1)

def VerificarNumberExtenso(mensagem, user):
    number = user.get_number()
    id = user.get_id()

    if mensagem.capitalize() == decidir_verificacao_number(int(number)):
                resposta = f'Você acertou!'
                user.set_ultimoComando('/Aprender')
                responder(id, resposta, [['Exibir Number', '/ExibirNumber'], ['Continuar','/OK']], 1)
    else:
        extenso = decidir_verificacao_number(int(number)).capitalize()
        resposta = f'Você errou! a traducao correta é: \n\n{extenso}'
        user.set_ultimoComando('/Aprender')
        responder(id, resposta, [['Exibir Number', '/ExibirNumber'], ['Continuar','/OK']], 1)
