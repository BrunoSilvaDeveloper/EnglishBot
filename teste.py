import random
from openpyxl import load_workbook

numero = random.randint(0, 99999)
print(numero)


def verificar_dezena(numero):
    planilha = load_workbook('Aprendendo.xlsx')
    aba_ativa = planilha.get_sheet_by_name('Numeros')

    if numero >= 0 and numero < 21:
        return aba_ativa[f'A{numero+1}'].value

    elif numero == 30:
        return f'{aba_ativa[f'A{22}'].value}'

    elif numero == 40:
        return f'{aba_ativa[f'A{23}'].value}'

    elif numero == 50:
        return f'{aba_ativa[f'A{24}'].value}'

    elif numero == 60:
        return f'{aba_ativa[f'A{25}'].value}'

    elif numero == 70:
        return f'{aba_ativa[f'A{26}'].value}'

    elif numero == 80:
        return f'{aba_ativa[f'A{27}'].value}'

    elif numero == 90:
        return f'{aba_ativa[f'A{28}'].value}'
        
    elif numero > 20 and numero < 30:
        return f'{aba_ativa[f'A{21}'].value}-{aba_ativa[f'A{numero-19}'].value}'

    elif numero > 30 and numero < 40:
        return f'{aba_ativa[f'A{22}'].value}-{aba_ativa[f'A{numero-29}'].value}'

    elif numero > 40 and numero < 50:
        return f'{aba_ativa[f'A{23}'].value}-{aba_ativa[f'A{numero-39}'].value}'

    elif numero > 50 and numero < 60:
        return f'{aba_ativa[f'A{24}'].value}-{aba_ativa[f'A{numero-49}'].value}'

    elif numero > 60 and numero < 70:
        return f'{aba_ativa[f'A{25}'].value}-{aba_ativa[f'A{numero-59}'].value}'

    elif numero > 70 and numero < 80:
        return f'{aba_ativa[f'A{26}'].value}-{aba_ativa[f'A{numero-69}'].value}'

    elif numero > 80 and numero < 90:
        return f'{aba_ativa[f'A{27}'].value}-{aba_ativa[f'A{numero-79}'].value}'

    elif numero > 90 and numero < 100:
        return f'{aba_ativa[f'A{28}'].value}-{aba_ativa[f'A{numero-89}'].value}'

def verificar_centena(numero):

    planilha = load_workbook('Aprendendo.xlsx')
    aba_ativa = planilha.get_sheet_by_name('Numeros')

    if numero == 100:
        return f'{aba_ativa[f'A2'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 200:
        return f'{aba_ativa[f'A3'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 300:
        return f'{aba_ativa[f'A4'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 400:
        return f'{aba_ativa[f'A5'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 500:
        return f'{aba_ativa[f'A6'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 600:
        return f'{aba_ativa[f'A7'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 700:
        return f'{aba_ativa[f'A8'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 800:
        return f'{aba_ativa[f'A9'].value}-{aba_ativa[f'A{29}'].value}'
    
    if numero == 900:
        return f'{aba_ativa[f'A10'].value}-{aba_ativa[f'A{29}'].value}'
        
    elif numero > 100 and numero < 200:
        return f'{aba_ativa[f'A2'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-100)}'

    elif numero > 200 and numero < 300:
        return f'{aba_ativa[f'A3'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-200)}'
    
    elif numero > 300 and numero < 400:
        return f'{aba_ativa[f'A4'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-300)}'

    elif numero > 400 and numero < 500:
        return f'{aba_ativa[f'A5'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-400)}'
    
    elif numero > 500 and numero < 600:
        return f'{aba_ativa[f'A6'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-500)}'
    
    elif numero > 600 and numero < 700:
        return f'{aba_ativa[f'A7'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-600)}'
    
    elif numero > 700 and numero < 800:
        return f'{aba_ativa[f'A8'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-700)}'

    elif numero > 800 and numero < 900:
        return f'{aba_ativa[f'A9'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-800)}'

    elif numero > 900 and numero < 1000:
        return f'{aba_ativa[f'A10'].value}-{aba_ativa[f'A{29}'].value} and {verificar_dezena(numero-900)}'

def verificar_milhar(numero):
    n = str(numero)
    if len(n) == 4:
        milhar = int(n[0])
        centena = int(n[-3])*100
        dezena = int(n[-2]+n[-1])
    elif len(n) == 5:
        milhar = int(n[0]+n[1])
        centena = int(n[-3])*100
        dezena = int(n[-2]+n[-1])

    if numero > 999 and numero < 100000:
        return f'{verificar_dezena(milhar)}-thousand and {verificar_centena(centena)} and {verificar_dezena(dezena)}'

def decidir_verificacao_number(numero):
    if numero < 100:
        print(verificar_dezena(numero))

    elif numero >=100 and numero < 1000:
        print(verificar_centena(numero))

    elif numero >=1000:
        print(verificar_milhar(numero))